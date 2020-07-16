from collections.abc import Iterable

from openpyxl import Workbook
from openpyxl import load_workbook
import tkinter as tk
import tkinter.filedialog as tkfiledialog
import tkinter.messagebox as tkmessagebox
import  os

# filePath=r'D:\转换\20200627-eHR_员工花名册报表与接口数据比对反馈.xlsx'
#filePath=r'D:\转换\工作簿1.xlsx'
root = tk.Tk().withdraw()
filePath = tkfiledialog.askopenfilename(title='请选择xlsx文件',filetypes=[("xlsx文件", "*.xlsx")])
print(f'filePath:{filePath}')
filePathDirPath=os.path.abspath(os.path.dirname(filePath))  #获取文件的上级目录
wb = load_workbook(filename=filePath,read_only=True)
print(wb.sheetnames)

def join2str(iter:Iterable,split='|'):
    """将数组,tuple等里面的元素拼接成字符串"""
    text = ''
    for v in iter:
        vv = str(v).strip().replace('\n','') #strip()去掉两端空格, 换行符
        if vv=='None':vv=''
        text += ( vv + split)
    if text.endswith(split): text=text[:len(text)-1]+'\n'
    print(text)
    return text

for sheet in wb:
    print(sheet.title)
    creatSH_FilePath=filePathDirPath +"\\"+ sheet.title + '.txt'   # 生成文件路径
    rows_values=sheet.values
    with open(creatSH_FilePath, mode='a',encoding='utf-8',newline='\n') as f:  # mode='a' 追加, 'w' 覆盖
        for rowNum,rowValue in enumerate(rows_values):
            # print(rowValue)
            if rowNum==0:continue   # 去掉Excel中的header
            rowValueStr=join2str(rowValue)   #拼接成字符串
            f.write(rowValueStr)   #写入文件
    print('=====================')
    print(f'生成文件路径: {creatSH_FilePath}')





exit_input=input('\n\n回车键退出!\n\n')
